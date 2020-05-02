import openpyxl

filename = '../../docs/2020.04.06.xlsx'
# Read the excel file(workbook)
wb = openpyxl.load_workbook(filename)

# Find sheet name -- no necessary if the workbook has only one sheet.

# Open the current sheet
sheet = wb.active

r = sheet.max_row
c = sheet.max_column
dr = 2       # the first row of the data
while dr <= r:
    dc = 5   # the first column of the data
    print('\n', 'row#', dr, ': ', end='')
    while dc <= c:
        print(sheet.cell(row=dr, column=dc).value, '\t', sep=',', end='', flush=True)
        dc += 1
    dr += 1

# Test
# print(sheet.max_row, sheet.max_column)
# print(sheet.cell(row=2, column=5).value)

# Close the excel file
wb.close()