import sys
import pymysql
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QMessageBox, QDesktopWidget, QFileDialog, QTextEdit

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        btnConn = QPushButton('DB Connection Test', self)
        btnConn.setCheckable(True)

        btnFile = QPushButton('File Open Dialog Test', self)

        self.textEdit = QTextEdit()

        vbox = QVBoxLayout()
        vbox.addWidget(btnConn)
        vbox.addWidget(btnFile)
        vbox.addWidget(self.textEdit)
        self.setLayout(vbox)

        btnConn.clicked.connect(self.dbConnect)
        btnFile.clicked.connect(self.showDialog)

        self.setWindowTitle('Test Application')
        # self.move(300, 300)
        self.resize(400, 200)
        # self.setGeometry(300, 300, 500, 200)
        self.center()
        self.show()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Message', 'Are you sure to quit?', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

    def dbConnect(self):
        try:
            conn = pymysql.connect(host='34.82.158.37', user='dev1', password='dev2020', db='myasset', charset='utf8')
            msg = 'DB Connection is successful'
            conn.close()
        except Exception as e:
            msg = 'DB Connection is failed'

        reply = QMessageBox().information(self, 'DB Connection', msg, QMessageBox.Ok)
        if reply == QMessageBox.Ok:
            self.show()

    def showDialog(self):
        fname = QFileDialog.getOpenFileName(self, 'Open file', './')
        if fname[0]:
            # f = open(fname[0], 'r')
            # with f:
            #     data = f.read()
            #     self.textEdit.setText(data)
            wb = openpyxl.load_workbook(fname[0])
            sheet = wb.active
            r = sheet.max_row
            c = sheet.max_column
            wb.close()
            textMessage = 'max row: ' + str(r) + '   ' + 'max column: ' + str(c)
            try:
                self.textEdit.setText(textMessage)
            except Exception as e:
                print(e)


if __name__ == '__main__':
   app = QApplication(sys.argv)
   ex = MyApp()
   sys.exit(app.exec_())